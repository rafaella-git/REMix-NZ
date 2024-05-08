import logging
from pathlib import Path

from pandas import DataFrame
from pandas import ExcelFile
from pandas import ExcelWriter
from pandas import read_excel

from remix.framework.api.instance import Instance
from remix.framework.schema.templates import default_dataframes

try:
    from openpyxl import load_workbook
except ImportError as e:
    logging.warning("To use the Excel backend, install openpyxl. Tip: use pip install remix.framework[excel]")
    load_workbook = None


class ExcelInstance(Instance):
    """Excel workbook based  REMix instance (experimental).

    .. note::

        This feature is experimental and probably unstable.
        Try exporting your excel workbook to csv files first.

    """
    def __init__(self, **kwargs) -> None:
        if load_workbook is None:
            raise ImportError("Please install the openpyxl library to use this feature. Tip: use pip install remix.framework[excel]")
        super().__init__(**kwargs)
        super().__init__(**kwargs)
        self.workbook_path = kwargs.get("workbook", None)
        if self.workbook_path is not None:
            if Path(self.workbook_path).exists():
                self.load_workbook_data()

    def load_workbook_data(self) -> DataFrame:
        """Assuming all the input data is in the same workbook."""
        dataframes = default_dataframes(index_names=True, column_names=True)
        workbook = ExcelFile(self.workbook_path)
        sheet_names = [s.lower() for s in workbook.sheet_names]
        for element in ["parameters", "timeseries", "maps"]:
            for param in dataframes[element].keys():
                if param.lower() in sheet_names:
                    file_dataframe = read_excel(workbook, sheet_name=param.lower())
                    index_names = list(dataframes[element][param].index.names)
                    # index_names = [idx for idx in index_names if idx.lower() != "timedata"]
                    file_dataframe = file_dataframe.set_index(index_names)
                    dataframes[element][param] = self._merge_dfs([dataframes[element][param], file_dataframe])
        self.assign_attributes(descriptor=dataframes)
        self.update_dataframes()
        self.ready = True

    def write_map_file(self, data, parent, name, extension, **kwargs):
        if extension == ".xlsx":
            if isinstance(data, list):
                data = self._merge_dfs(data)
            data = data.reset_index()
            data.to_excel(self.workbook, sheet_name=name, index=False)
        else:
            super().write_map_file(data, parent, name, extension, **kwargs)

    def write_parameter_file(self, data, parent, name, extension, **kwargs):
        if extension == ".xlsx":
            if isinstance(data, list):
                data = self._merge_dfs(data)
            data = data.reset_index()
            data.to_excel(self.workbook, sheet_name=name, index=False)
        else:
            super().write_parameter_file(data, parent, name, extension, **kwargs)

    def write_set_file(self, data, parent, name, extension, **kwargs):
        if extension == ".xlsx":
            pass
        else:
            super().write_set_file(data, parent, name, extension, **kwargs)

    def write_profile_file(self, data, parent, name, extension, **kwargs):
        if extension == ".xlsx":
            if isinstance(data, list):
                data = self._merge_dfs(data)
            data = data.reset_index()
            data.to_excel(self.workbook, sheet_name=name, index=False)
        else:
            super().write_profile_file(data, parent, name, extension, **kwargs)

    def write(self, inspect=False, fileformat="csv", infer=True):
        """This write method exports your input excel file as csv like a normal instance.
        if excel is given as fileformat it will export a workbook instead.

        Parameters
        ----------
        inspect : bool
            If true it will return a dictionary with the output files. Defaults to False.
        fileformat : str
            "xlsx" will create an excel workbook. Defaults to "csv".
        infer : bool
            If true, it will infer set information from the input DataFrames. Defaults to True.

        Returns
        -------
        pandas.DataFrame
            In case inspect is True, else returns NoneType.
        """
        if fileformat.lower() == "excel":
            fileformat == "xlsx"
        if fileformat == "xlsx":
            if self.workbook_path is not None:
                if Path(self.workbook_path).exists():
                    workbook = load_workbook(self.workbook_path)
                    self.workbook = ExcelWriter(self.workbook_path, engine='openpyxl')
                    self.workbook.book = workbook
                else:
                    self.workbook = ExcelWriter(self.workbook_path, engine='openpyxl')
            else:
                self.workbook = ExcelWriter(Path(self.datadir).joinpath("remix.xlsx"), engine='openpyxl')
        dataframes = super().write(inspect=inspect, fileformat=fileformat, infer=infer)
        if fileformat == "xlsx":
            self.workbook.close()
        return dataframes

    @classmethod
    def from_dataframes(
        cls, dataframes: dict, version="latest", name="default", **kwargs
    ):
        instance = super().from_dataframes(
            dataframes=dataframes, version=version, name=name, index_names=True, **kwargs
        )
        return instance
